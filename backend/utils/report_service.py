import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from models import Customer, Quotation, db
from sqlalchemy import func
import io

def export_customers_excel(customers):
    """将客户数据导出为 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '客户列表'
    
    # 设置表头
    headers = ['公司名称', '国家', '城市', '邮箱', '电话', 'WhatsApp', '客户类型', 
               '行业', '等级', 'AI 评分', '状态', '来源', '创建日期', '最后联系']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 添加数据
    for row, customer in enumerate(customers, 2):
        ws.cell(row=row, column=1).value = customer.name
        ws.cell(row=row, column=2).value = customer.country
        ws.cell(row=row, column=3).value = customer.city
        ws.cell(row=row, column=4).value = customer.email
        ws.cell(row=row, column=5).value = customer.phone
        ws.cell(row=row, column=6).value = customer.whatsapp
        ws.cell(row=row, column=7).value = customer.company_type
        ws.cell(row=row, column=8).value = customer.industry
        ws.cell(row=row, column=9).value = customer.customer_grade
        ws.cell(row=row, column=10).value = customer.ai_score
        ws.cell(row=row, column=11).value = customer.status
        ws.cell(row=row, column=12).value = customer.source
        ws.cell(row=row, column=13).value = customer.created_at.strftime('%Y-%m-%d') if customer.created_at else ''
        ws.cell(row=row, column=14).value = customer.last_contact.strftime('%Y-%m-%d') if customer.last_contact else ''
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 返回二进制流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def export_quotations_excel(quotations):
    """将报价数据导出为 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '报价列表'
    
    headers = ['客户名称', '产品', '数量', '单价', '总价', '货币', '状态', 
               '创建日期', '发送日期', '回复日期', '有效期(天)']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row, quotation in enumerate(quotations, 2):
        ws.cell(row=row, column=1).value = quotation.customer.name
        ws.cell(row=row, column=2).value = quotation.product_type
        ws.cell(row=row, column=3).value = quotation.quantity
        ws.cell(row=row, column=4).value = quotation.unit_price
        ws.cell(row=row, column=5).value = quotation.total_amount
        ws.cell(row=row, column=6).value = quotation.currency
        ws.cell(row=row, column=7).value = quotation.status
        ws.cell(row=row, column=8).value = quotation.created_at.strftime('%Y-%m-%d')
        ws.cell(row=row, column=9).value = quotation.sent_at.strftime('%Y-%m-%d') if quotation.sent_at else ''
        ws.cell(row=row, column=10).value = quotation.responded_at.strftime('%Y-%m-%d') if quotation.responded_at else ''
        ws.cell(row=row, column=11).value = quotation.validity_days
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def export_sales_report(user_id=None, start_date=None, end_date=None):
    """生成销售报告"""
    from datetime import datetime, timedelta
    
    # 默认时间范围：最近 30 天
    if not end_date:
        end_date = datetime.now()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    
    if not start_date:
        start_date = end_date - timedelta(days=30)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    
    query = Customer.query
    if user_id:
        query = query.filter(Customer.owner_id == user_id)
    
    query = query.filter(
        Customer.created_at >= start_date,
        Customer.created_at <= end_date
    )
    
    total_customers = query.count()
    total_quotations = db.session.query(func.count(Quotation.id)).filter(
        Quotation.created_at >= start_date,
        Quotation.created_at <= end_date,
        Quotation.created_by == user_id if user_id else True
    ).scalar()
    
    total_amount = db.session.query(func.sum(Quotation.total_amount)).filter(
        Quotation.created_at >= start_date,
        Quotation.created_at <= end_date,
        Quotation.created_by == user_id if user_id else True
    ).scalar() or 0
    
    accepted = db.session.query(func.count(Quotation.id)).filter(
        Quotation.status == 'accepted',
        Quotation.created_at >= start_date,
        Quotation.created_at <= end_date,
        Quotation.created_by == user_id if user_id else True
    ).scalar()
    
    return {
        'period': f'{start_date.strftime("%Y-%m-%d")} 到 {end_date.strftime("%Y-%m-%d")}',
        'total_customers': total_customers,
        'total_quotations': total_quotations,
        'total_amount': float(total_amount),
        'accepted_quotations': accepted,
        'conversion_rate': round(accepted / total_quotations * 100, 2) if total_quotations > 0 else 0
    }
